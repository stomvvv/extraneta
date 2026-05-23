"""
Unit tests for the Yandex Travel parser using real file structure.
"""

import io
import pytest
from decimal import Decimal
from datetime import date

import openpyxl

from app.services.parsers.yandex import YandexParser
from app.models.booking import OTASource, PaymentStatus, BookingStatus


def _make_yandex_xlsx(rows: list[dict]) -> bytes:
    """Build a minimal Yandex-format xlsx in memory."""
    wb = openpyxl.Workbook()
    ws = wb.active

    # Row 0: payment ID
    ws.cell(1, 1, "1803097100")
    # Row 1: empty
    # Row 2 (index=2): headers
    headers = [
        "Объект размещения",
        "№ Яндекс",
        "Имя гостя",
        "Дата брони",
        "Заезд",
        "Выезд",
        "Тип перечисления",
        "Оплатил гость, ₽",
        "Перечислено отелю, ₽",
        "Вознаграждение Яндекса с учетом скидки по совместной акции",
        "Ставка вознаграждения",
    ]
    for ci, h in enumerate(headers, 1):
        ws.cell(3, ci, h)

    for ri, row in enumerate(rows, 4):
        ws.cell(ri, 1, row.get("hotel"))
        ws.cell(ri, 2, row.get("booking_id"))
        ws.cell(ri, 3, row.get("guest"))
        ws.cell(ri, 4, row.get("booking_date"))
        ws.cell(ri, 5, row.get("check_in"))
        ws.cell(ri, 6, row.get("check_out"))
        ws.cell(ri, 7, row.get("type", "Оплата"))
        ws.cell(ri, 8, row.get("gross"))
        ws.cell(ri, 9, row.get("net"))
        ws.cell(ri, 10, row.get("commission"))
        ws.cell(ri, 11, row.get("rate", "15.00%"))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


class TestYandexParser:
    def setup_method(self):
        self.parser = YandexParser()

    def test_parse_single_booking(self):
        xlsx = _make_yandex_xlsx([{
            "hotel": "Shato City",
            "booking_id": "YA-0579-2896-4276",
            "guest": "Иван Иванов",
            "booking_date": "2024-11-12",
            "check_in": "2025-01-04",
            "check_out": "2025-01-06",
            "gross": 15624,
            "net": 13280.4,
            "commission": 2343.6,
            "rate": "15.00%",
        }])
        result = self.parser.parse(xlsx, "test.xlsx")

        assert not result.errors
        assert len(result.bookings) == 1
        b = result.bookings[0]
        assert b.booking_id_ota == "YA-0579-2896-4276"
        assert b.guest_name == "Иван Иванов"
        assert b.check_in == date(2025, 1, 4)
        assert b.check_out == date(2025, 1, 6)
        assert b.nights == 2
        assert b.gross_amount == Decimal("15624")
        assert b.net_amount == Decimal("13280.4")
        assert b.ota_commission_amount == Decimal("2343.6")
        assert b.ota_commission_rate == Decimal("15.00")
        assert b.payment_status == PaymentStatus.paid
        assert b.booking_status == BookingStatus.confirmed
        assert b.source_ota == OTASource.yandex
        assert b.hotel_name_from_file == "Shato City"
        assert b.currency == "RUB"

    def test_parse_refund(self):
        xlsx = _make_yandex_xlsx([{
            "hotel": "Shato Plaza",
            "booking_id": "YA-8675-7262-3623",
            "guest": "Максим Бриллиант",
            "check_in": "2025-01-07",
            "check_out": "2025-01-08",
            "type": "Возврат",
            "gross": -2750,
            "net": -2337.5,
            "commission": -412.5,
            "rate": "15.00%",
        }])
        result = self.parser.parse(xlsx, "test.xlsx")
        assert len(result.bookings) == 1
        b = result.bookings[0]
        assert b.payment_status == PaymentStatus.refunded
        assert b.booking_status == BookingStatus.cancelled
        assert b.gross_amount < 0

    def test_skip_non_ya_rows(self):
        xlsx = _make_yandex_xlsx([
            {"hotel": "", "booking_id": "", "guest": "", "check_in": "", "check_out": "", "gross": 0, "net": 0, "commission": 0},
            {"hotel": "Shato", "booking_id": "YA-0001-0002-0003", "guest": "Гость", "check_in": "2025-01-01", "check_out": "2025-01-02", "gross": 5000, "net": 4250, "commission": 750},
        ])
        result = self.parser.parse(xlsx, "test.xlsx")
        assert len(result.bookings) == 1
        assert result.skipped >= 1

    def test_multi_property(self):
        xlsx = _make_yandex_xlsx([
            {"hotel": "Shato City", "booking_id": "YA-AAA-BBB-CCC", "guest": "Гость 1", "check_in": "2025-01-01", "check_out": "2025-01-02", "gross": 5000, "net": 4250, "commission": 750},
            {"hotel": "Shato Plaza", "booking_id": "YA-DDD-EEE-FFF", "guest": "Гость 2", "check_in": "2025-01-01", "check_out": "2025-01-03", "gross": 8000, "net": 6800, "commission": 1200},
        ])
        result = self.parser.parse(xlsx, "test.xlsx")
        assert len(result.bookings) == 2
        hotels = {b.hotel_name_from_file for b in result.bookings}
        assert "Shato City" in hotels
        assert "Shato Plaza" in hotels

    def test_nights_calculation(self):
        xlsx = _make_yandex_xlsx([{
            "hotel": "Test", "booking_id": "YA-111-222-333", "guest": "Гость",
            "check_in": "2025-01-01", "check_out": "2025-01-05",
            "gross": 10000, "net": 8500, "commission": 1500, "rate": "15.00%",
        }])
        result = self.parser.parse(xlsx, "test.xlsx")
        assert result.bookings[0].nights == 4

    def test_commission_rate_parsing(self):
        """Commission rate '15.00%' should parse to Decimal('15.00'), not 0.15."""
        xlsx = _make_yandex_xlsx([{
            "hotel": "Test", "booking_id": "YA-000-000-001", "guest": "Гость",
            "check_in": "2025-01-01", "check_out": "2025-01-02",
            "gross": 7812, "net": 6640.2, "commission": 1171.8, "rate": "15.00%",
        }])
        result = self.parser.parse(xlsx, "test.xlsx")
        assert result.bookings[0].ota_commission_rate == Decimal("15.00")
