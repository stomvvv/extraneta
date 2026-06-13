import io
from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import and_
from datetime import date, timedelta
from typing import Optional
from database import get_db
from models.booking import Booking
from models.hotel import Hotel

router = APIRouter()


def _get_bookings(db, hotel_id, date_from, date_to):
    filters = []
    if hotel_id:
        filters.append(Booking.hotel_id == hotel_id)
    if date_from:
        filters.append(Booking.check_in >= date_from)
    if date_to:
        filters.append(Booking.check_in <= date_to)
    q = db.query(Booking)
    if filters:
        q = q.filter(and_(*filters))
    return q.order_by(Booking.check_in.desc()).all()


@router.get("/reports/excel")
def download_excel(
    hotel_id: Optional[int] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    db: Session = Depends(get_db),
):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    if not hotel_id:
        first = db.query(Hotel).order_by(Hotel.id).first()
        hotel_id = first.id if first else None

    d_from = date.fromisoformat(date_from) if date_from else date.today() - timedelta(days=365)
    d_to = date.fromisoformat(date_to) if date_to else date.today()

    bookings = _get_bookings(db, hotel_id, d_from, d_to)

    wb = openpyxl.Workbook()

    # Sheet 1: Summary by OTA
    ws1 = wb.active
    ws1.title = "Сводка по OTA"
    ws1.append(["OTA", "Бронирований", "Валовая выручка", "Комиссия ₽", "Комиссия %", "Нетто"])
    for cell in ws1[1]:
        cell.font = Font(bold=True)

    agg: dict[str, dict] = {}
    for b in bookings:
        if b.booking_status != "confirmed":
            continue
        ota = b.source_ota
        if ota not in agg:
            agg[ota] = {"count": 0, "gross": 0.0, "commission": 0.0}
        agg[ota]["count"] += 1
        agg[ota]["gross"] += float(b.gross_amount)
        agg[ota]["commission"] += float(b.ota_commission_amount)

    for ota, v in sorted(agg.items()):
        rate = v["commission"] / v["gross"] * 100 if v["gross"] else 0
        ws1.append([ota, v["count"], round(v["gross"], 2), round(v["commission"], 2),
                    round(rate, 2), round(v["gross"] - v["commission"], 2)])

    # Sheet 2: All bookings
    ws2 = wb.create_sheet("Все бронирования")
    ws2.append(["OTA", "ID брони", "Гость", "Заезд", "Выезд", "Ночей",
                "Валовая", "Комиссия %", "Комиссия ₽", "Нетто", "Статус", "Аномалия"])
    for cell in ws2[1]:
        cell.font = Font(bold=True)
    for b in bookings:
        ws2.append([b.source_ota, b.booking_id_ota, b.guest_name,
                    str(b.check_in), str(b.check_out), b.nights,
                    float(b.gross_amount), float(b.ota_commission_rate),
                    float(b.ota_commission_amount), float(b.net_amount),
                    b.booking_status, "Да" if b.has_anomaly else ""])

    # Sheet 3: Anomalies
    ws3 = wb.create_sheet("Аномалии")
    ws3.append(["OTA", "ID брони", "Гость", "Заезд", "Валовая", "Комиссия %", "Причина"])
    for cell in ws3[1]:
        cell.font = Font(bold=True)
    for b in bookings:
        if b.has_anomaly:
            ws3.append([b.source_ota, b.booking_id_ota, b.guest_name,
                        str(b.check_in), float(b.gross_amount),
                        float(b.ota_commission_rate), b.anomaly_reason])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"extraneta_report_{d_from}_{d_to}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/reports/pdf")
def download_pdf(
    hotel_id: Optional[int] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    db: Session = Depends(get_db),
):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.units import cm
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    if not hotel_id:
        first = db.query(Hotel).order_by(Hotel.id).first()
        hotel_id = first.id if first else None

    d_from = date.fromisoformat(date_from) if date_from else date.today() - timedelta(days=365)
    d_to = date.fromisoformat(date_to) if date_to else date.today()

    bookings = _get_bookings(db, hotel_id, d_from, d_to)
    confirmed = [b for b in bookings if b.booking_status == "confirmed"]

    hotel = db.get(Hotel, hotel_id) if hotel_id else None
    hotel_name = hotel.name if hotel else "Гостиница"

    gross = sum(float(b.gross_amount) for b in confirmed)
    commission = sum(float(b.ota_commission_amount) for b in confirmed)
    net = gross - commission

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, rightMargin=2*cm, leftMargin=2*cm,
                            topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    story = []

    story.append(Paragraph(f"ExtranEta — Управленческий отчёт", styles["Title"]))
    story.append(Paragraph(f"{hotel_name} | {d_from} — {d_to}", styles["Normal"]))
    story.append(Spacer(1, 0.5*cm))

    # KPI table
    kpi_data = [
        ["Показатель", "Значение"],
        ["Валовая выручка", f"{gross:,.2f} руб."],
        ["Комиссии OTA", f"{commission:,.2f} руб."],
        ["Чистая выручка", f"{net:,.2f} руб."],
        ["Средний % комиссии", f"{commission/gross*100:.1f}%" if gross else "0%"],
        ["Всего бронирований", str(len(bookings))],
        ["Подтверждённых", str(len(confirmed))],
        ["Аномалий", str(sum(1 for b in bookings if b.has_anomaly))],
    ]
    t = Table(kpi_data, colWidths=[10*cm, 7*cm])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E3A5F")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8f9fa")]),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
    ]))
    story.append(t)
    story.append(Spacer(1, 0.5*cm))

    # OTA breakdown
    agg: dict[str, dict] = {}
    for b in confirmed:
        ota = b.source_ota
        if ota not in agg:
            agg[ota] = {"count": 0, "gross": 0.0, "commission": 0.0}
        agg[ota]["count"] += 1
        agg[ota]["gross"] += float(b.gross_amount)
        agg[ota]["commission"] += float(b.ota_commission_amount)

    if agg:
        story.append(Paragraph("Разбивка по OTA-каналам", styles["Heading2"]))
        rows = [["OTA", "Броней", "Выручка", "Комиссия", "%"]]
        for ota, v in sorted(agg.items(), key=lambda x: -x[1]["gross"]):
            rate = v["commission"] / v["gross"] * 100 if v["gross"] else 0
            rows.append([ota, str(v["count"]), f"{v['gross']:,.0f}", f"{v['commission']:,.0f}", f"{rate:.1f}%"])
        t2 = Table(rows, colWidths=[5*cm, 2.5*cm, 4*cm, 3.5*cm, 2*cm])
        t2.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#6366f1")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8f9fa")]),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
        ]))
        story.append(t2)

    doc.build(story)
    buf.seek(0)
    filename = f"extraneta_report_{d_from}_{d_to}.pdf"
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
