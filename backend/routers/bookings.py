import io
from fastapi import APIRouter, Depends, Query, Response
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import and_
from datetime import date
from typing import Optional
from database import get_db
from models.booking import Booking
from models.hotel import Hotel

router = APIRouter()


@router.get("/bookings")
def list_bookings(
    hotel_id: Optional[int] = None,
    page: int = 1,
    limit: int = 50,
    ota: Optional[str] = None,
    status: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    is_anomaly: Optional[bool] = None,
    guest: Optional[str] = None,
    db: Session = Depends(get_db),
):
    if not hotel_id:
        first = db.query(Hotel).order_by(Hotel.id).first()
        hotel_id = first.id if first else None

    filters = []
    if hotel_id:
        filters.append(Booking.hotel_id == hotel_id)
    if ota:
        filters.append(Booking.source_ota == ota)
    if status:
        filters.append(Booking.booking_status == status)
    if is_anomaly is not None:
        filters.append(Booking.has_anomaly == is_anomaly)
    if date_from:
        try:
            filters.append(Booking.check_in >= date.fromisoformat(date_from))
        except Exception:
            pass
    if date_to:
        try:
            filters.append(Booking.check_in <= date.fromisoformat(date_to))
        except Exception:
            pass
    if guest:
        filters.append(Booking.guest_name.ilike(f"%{guest}%"))

    q = db.query(Booking)
    if filters:
        q = q.filter(and_(*filters))

    total = q.count()
    items = q.order_by(Booking.check_in.desc()).offset((page - 1) * limit).limit(limit).all()

    return {
        "total": total,
        "page": page,
        "limit": limit,
        "pages": (total + limit - 1) // limit,
        "items": [_booking_dict(b) for b in items],
    }


@router.get("/bookings/export")
def export_bookings(
    hotel_id: Optional[int] = None,
    ota: Optional[str] = None,
    status: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    db: Session = Depends(get_db),
):
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    if not hotel_id:
        first = db.query(Hotel).order_by(Hotel.id).first()
        hotel_id = first.id if first else None

    filters = []
    if hotel_id:
        filters.append(Booking.hotel_id == hotel_id)
    if ota:
        filters.append(Booking.source_ota == ota)
    if status:
        filters.append(Booking.booking_status == status)
    if date_from:
        try:
            filters.append(Booking.check_in >= date.fromisoformat(date_from))
        except Exception:
            pass
    if date_to:
        try:
            filters.append(Booking.check_in <= date.fromisoformat(date_to))
        except Exception:
            pass

    q = db.query(Booking)
    if filters:
        q = q.filter(and_(*filters))
    bookings = q.order_by(Booking.check_in.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Бронирования"
    headers = ["OTA", "ID брони", "Гость", "Тип номера", "Заезд", "Выезд", "Ночей",
               "Валовая", "Комиссия %", "Комиссия ₽", "Нетто", "Статус", "Аномалия"]
    ws.append(headers)
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font

    for b in bookings:
        ws.append([
            b.source_ota, b.booking_id_ota, b.guest_name, b.room_type,
            str(b.check_in), str(b.check_out), b.nights,
            float(b.gross_amount), float(b.ota_commission_rate),
            float(b.ota_commission_amount), float(b.net_amount),
            b.booking_status, "Да" if b.has_anomaly else "",
        ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=bookings.xlsx"},
    )


def _booking_dict(b: Booking) -> dict:
    return {
        "id": b.id,
        "hotel_id": b.hotel_id,
        "upload_id": b.upload_id,
        "source_ota": b.source_ota,
        "booking_id_ota": b.booking_id_ota,
        "guest_name": b.guest_name,
        "room_type": b.room_type,
        "check_in": str(b.check_in),
        "check_out": str(b.check_out),
        "nights": b.nights,
        "gross_amount": float(b.gross_amount),
        "ota_commission_rate": float(b.ota_commission_rate),
        "ota_commission_amount": float(b.ota_commission_amount),
        "net_amount": float(b.net_amount),
        "currency": b.currency,
        "payment_status": b.payment_status,
        "booking_status": b.booking_status,
        "has_anomaly": b.has_anomaly,
        "anomaly_reason": b.anomaly_reason,
        "created_at": str(b.created_at),
    }
